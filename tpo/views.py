from django.shortcuts import render, redirect
from django.contrib.auth.hashers import make_password, check_password
from django.contrib import messages
from .models import Tpo, Job
# Create your views here.
# def tpo_dashboard(request):
#     if "tpo_id" not in request.session:
#         messages.error(request, "You need to log in first.")
#         return redirect("tpo_login")
#     tpo = Tpo.objects.get(id=request.session["tpo_id"])

#     return render(request, 'tpo_dashboard.html', { 'tpo': tpo })


from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Job
from users.models import JobApplication, User
from tpo.models import Tpo
def tpo_dashboard(request):
    if "tpo_id" not in request.session:
        messages.error(request, "You need to log in first.")
        return redirect("tpo_login")
    
    tpo = Tpo.objects.get(id=request.session["tpo_id"])
    
    # Total counts and statistics
    total_jobs = Job.objects.count()
    total_applications = JobApplication.objects.count()
    total_candidates = User.objects.count()
    shortlisted = JobApplication.objects.filter(status="shortlisted").count()
    ongoing_interviews = JobApplication.objects.filter(status="interviewing").count()
    placed = JobApplication.objects.filter(status="placed").count()
    rejected = JobApplication.objects.filter(status="rejected").count()

    job_details = Job.objects.all()

    # Create a dictionary with job-wise application counts
    job_application_counts = {
        job.id: {
            "total": JobApplication.objects.filter(job=job).count(),
            "shortlisted": JobApplication.objects.filter(job=job, status="shortlisted").count(),
            "interviewing": JobApplication.objects.filter(job=job, status="interviewing").count(),
            "placed": JobApplication.objects.filter(job=job, status="placed").count(),
            "rejected": JobApplication.objects.filter(job=job, status="rejected").count(),
        }
        for job in job_details
    }

    # Debugging: print the job_application_counts to ensure it's being populated
    print("Job Application Counts:", job_application_counts)

    context = {
        'tpo': tpo,
        'total_jobs': total_jobs,
        'total_applications': total_applications,
        'total_candidates': total_candidates,
        'shortlisted': shortlisted,
        'ongoing_interviews': ongoing_interviews,
        'placed': placed,
        'rejected': rejected,
        'job_details': job_details,
        'job_application_counts': job_application_counts,
    }

    return render(request, 'tpo_dashboard.html', context)

# import openpyxl
# from django.http import HttpResponse
# from django.utils.dateparse import parse_date
# from datetime import datetime
# from users.models import Job  # or your model name

# def export_to_excel(request):
#     from_date = request.GET.get('from_date')
#     to_date = request.GET.get('to_date')

#     if from_date and to_date:
#         from_date = parse_date(from_date)
#         to_date = parse_date(to_date)

#         jobs = Job.objects.filter(created_at__range=(from_date, to_date))
#     else:
#         jobs = Job.objects.all()

#     # Create workbook
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Job Listings"

#     # Add headers
#     ws.append(['Job Title', 'Company', 'Location', 'Status', 'Created At'])

#     for job in jobs:
#         ws.append([
#             job.role,
#             job.company,
#             job.job_location,
#             'Active' if job.is_active else 'Inactive',
#             job.created_at.strftime("%Y-%m-%d")
#         ])

#     # Prepare response
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     )
#     response['Content-Disposition'] = f'attachment; filename=jobs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
#     wb.save(response)
#     return response

# import openpyxl
# from django.http import HttpResponse
# from django.utils.dateparse import parse_date
# from django.utils.timezone import make_aware
# from datetime import datetime
# from users.models import Job  # Make sure you're importing the correct model

# def export_to_excel(request):
#     # Get from_date and to_date from the GET request
#     from_date_str = request.GET.get('from_date')
#     to_date_str = request.GET.get('to_date')

#     # Check if both dates are provided
#     if from_date_str and to_date_str:
#         # Parse the date strings into datetime objects and make them timezone-aware
#         from_date = make_aware(datetime.strptime(from_date_str, '%Y-%m-%d'))
#         to_date = make_aware(datetime.strptime(to_date_str, '%Y-%m-%d'))
        
#         # Filter the jobs by the date range
#         jobs = Job.objects.filter(created_at__range=(from_date, to_date))
#     else:
#         # If no dates are provided, fetch all jobs
#         jobs = Job.objects.all()

#     # Create a new Excel workbook and set the sheet title
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Job Listings"

#     # Add headers to the Excel sheet
#     ws.append(['Job Title', 'Company', 'Location', 'Status', 'Created At'])

#     # Add job data to the Excel sheet
#     for job in jobs:
#         ws.append([
#             job.role,
#             job.company,
#             job.job_location,
#             'Active' if job.is_active else 'Inactive',
#             job.created_at.strftime("%Y-%m-%d")  # Formatting created_at as a string
#         ])

#     # Prepare the HTTP response with the Excel file as an attachment
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     )
#     response['Content-Disposition'] = f'attachment; filename=jobs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

#     # Save the workbook to the response, which will trigger the download
#     wb.save(response)
#     return response

import openpyxl
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from django.utils.timezone import make_aware
from datetime import datetime
from users.models import Job  # Adjust import if necessary

def export_to_excel(request):
    if "tpo_id" not in request.session:
        messages.error(request, "You need to log in first.")
        return redirect("tpo_login")
    
    tpo = Tpo.objects.get(id=request.session["tpo_id"])
    # Get from_date and to_date from the GET request
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')

    # Check if both dates are provided
    if from_date_str and to_date_str:
        # Parse the date strings into datetime objects and make them timezone-aware
        from_date = make_aware(datetime.strptime(from_date_str, '%Y-%m-%d'))
        to_date = make_aware(datetime.strptime(to_date_str, '%Y-%m-%d'))

        # Debugging: Log the filtered date range
        # print(f"Filtering jobs from {from_date} to {to_date}")

        # Filter the jobs by the date range
        jobs = Job.objects.filter(created_at__range=(from_date, to_date))

        # Debugging: log the number of jobs retrieved
    #     print(f"Filtered Jobs Count: {jobs.count()}")
    #     print("Jobs:", jobs)
    # else:
        # If no dates are provided, fetch all jobs
        jobs = Job.objects.all()

    # Create a new Excel workbook and set the sheet title
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Listings"

    # Add headers to the Excel sheet
    ws.append(['Job Title', 'Company', 'Location', 'Status', 'Created At'])

    # Add job data to the Excel sheet if there are jobs
    if jobs.exists():
        for job in jobs:
            ws.append([
                job.role,
                job.company,
                job.job_location,
                'Active' if job.is_active else 'Inactive',
                job.created_at.strftime("%Y-%m-%d")  # Formatting created_at as a string
            ])
    else:
        # No jobs found in the filtered range, notify user
        ws.append(['No jobs found in the selected date range.'])

    # Prepare the HTTP response with the Excel file as an attachment
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename=jobs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    # Save the workbook to the response, which will trigger the download
    wb.save(response)
    return response




def tpo_registration(request):
    if request.method == 'POST':
        full_name = request.POST['full_name']
        employee_id = request.POST['employee_id']
        email = request.POST['email']
        mobile_number = request.POST['mobile']
        password = request.POST['password']
        confirm_password = request.POST['confirm_password']

        # Password matching check
        if password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return render(request, 'registration.html', {
                'full_name': full_name,
                'employee_id': employee_id,
                'email': email,
                'mobile_number': mobile_number,
                
            })

        # Check if mobile number already exists
        if Tpo.objects.filter(employee_id=employee_id).exists():
            messages.error(request, "Employee id already registered.")
            return render(request, 'tpo_registration.html', {
                'full_name': full_name,
                'employee_id': employee_id,
                'email': email,
                'mobile_number': mobile_number,
                
            })

        # Hash password before storing
        hashed_password = make_password(password)

        # Create and save the vendor to the database
        tpo = Tpo(
            full_name=full_name,
            employee_id=employee_id,
            email=email,
            mobile_number=mobile_number,
        
            password=hashed_password
        )
        tpo.save()

        messages.success(request, "Registration successful! You can now log in.")
        return redirect('tpo_login')  # Redirect to login page after successful registration
    return render(request, 'tpo_registration.html')

def tpo_login(request):
    if request.method == "POST":
        employee_id = request.POST["employee_id"]
        password = request.POST["password"]

        try:
            tpo = Tpo.objects.get(employee_id=employee_id)

            if not tpo.is_approved:
                messages.error(request, "Your account is not approved by the admin.")
                return render(request, "tpo_login.html")
            
            if check_password(password, tpo.password):
                request.session["tpo_id"] = tpo.id
                print(tpo.id)
                request.session["tpo_name"] = tpo.full_name
                print(request.session["tpo_name"])
                messages.success(request, "Login successful.")
                return redirect("tpo_dashboard")  # Redirect vendors only

            else:
                messages.error(request, "Invalid password.")
                return render(request, "tpo_login.html")

        except Tpo.DoesNotExist:
            messages.error(request, "Tpo does not exist.")
            return render(request, "tpo_login.html")
    return render(request, 'tpo_login.html')

def tpo_logout(request):
    if "tpo_id" not in request.session:
        messages.error(request, "You need to log in first.")
        return redirect("tpo_login")
    
    tpo = Tpo.objects.get(id=request.session["tpo_id"])
    
    try:
        del request.session["tpo_id"]
        del request.session["tpo_name"]
        messages.success(request, "Logout successful.")
        return redirect('tpo_login')
    except KeyError:
        messages.error(request, "You are not logged in.")
        
        return redirect('tpo_login')

def tpo_update_profile(request):
    if "tpo_id" not in request.session:
        messages.error(request, "You need to log in first.")
        return redirect("tpo_login")
    
    tpo = Tpo.objects.get(id=request.session.get('tpo_id'))  # assuming session holds vendor_id

    if request.method == 'POST':
        full_name = request.POST.get('full_name')
        employee_id = request.POST.get('employee_id')
        email = request.POST.get('email')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        mobile = request.POST.get('mobile')

        if password != confirm_password:
            messages.error(request, "Passwords do not match.")
        else:
            tpo.full_name = full_name
            tpo.employee_id = employee_id
            tpo.email = email
            tpo.mobile_number = mobile
            tpo.password = password
            if password:
                tpo.password = make_password(password)
                tpo.save()
            messages.success(request, "Profile updated successfully.")
            request.session.update({"vendor_name": full_name})  # Update session name

    context = {
        'full_name': tpo.full_name,
        'employee_id': tpo.employee_id,
        'email': tpo.email,
        'mobile_number': tpo.mobile_number,
        
        
    }
    return render(request, 'tpo_account.html', context)


def add_job(request):
    if "tpo_id" not in request.session:
        messages.error(request, "You need to log in first.")
        return redirect("tpo_login")
    
    tpo = Tpo.objects.get(id=request.session["tpo_id"])
    tpo = Tpo.objects.get(id=request.session.get('tpo_id'))  # assuming session holds tpo_id
    if request.method == "POST":
        role = request.POST.get('role')
        company_name = request.POST.get('company_name')
        package = request.POST.get('salary')
        last_date = request.POST.get('last_date_to_apply')
        eligibility = request.POST.get('eligibility')
        job_description = request.POST.get('job_description')
        company_website = request.POST.get("company_website")
        job_location = request.POST.get("job_location")

        Job.objects.create(
            role=role,
            company=company_name,
            salary_package=package,
            last_date_to_apply=last_date,
            eligibility=eligibility,
            job_description=job_description,
            company_website=company_website,
            job_location=job_location,
            tpo=tpo  # Link the job to the logged-in tpo
        )
        messages.success(request, "Job added successfully.")
        return redirect('add-job')  # or redirect to a success page

    return render(request, 'add_job.html')

# views.py
from django.shortcuts import get_object_or_404, redirect
from .models import Job

def toggle_job_status(request, job_id):
    if "tpo_id" not in request.session:
        messages.error(request, "You need to log in first.")
        return redirect("tpo_login")
    
    tpo = Tpo.objects.get(id=request.session["tpo_id"])
    job = get_object_or_404(Job, id=job_id)
    job.is_active = not job.is_active
    job.save()
    return redirect('manage-job')  # or wherever you list jobs


def manage_job(request):
    if "tpo_id" not in request.session:
        messages.error(request, "You need to log in first.")
        return redirect("tpo_login")
    
    tpo = Tpo.objects.get(id=request.session["tpo_id"])

    tpo_id = request.session.get("tpo_id")

    if tpo_id is None:
        return redirect("tpo_login")  # Redirect if tpo is not logged in

    # Fetch jobs belonging to the logged-in tpo
    Jobs = Job.objects.all().order_by("-id")
    print("you are managing a job " ,Jobs)

    return render(request, "manage_job.html", {"jobs": Jobs})

def update_job(request, id):
    if "tpo_id" not in request.session:
        messages.error(request, "You need to log in first.")
        return redirect("tpo_login")
    
    tpo = Tpo.objects.get(id=request.session["tpo_id"])
    Jobs = get_object_or_404(Job, id=id)
    print("you are updatating a job " ,id)
    
    if request.method == "POST":
        print("updatting a job")
        
        Jobs.company = request.POST["company_name"]
        Jobs.role = request.POST["role"]
        Jobs.salary_package = request.POST["salary_package"]
        Jobs.last_date_to_apply = request.POST["last_date_to_apply"]
        Jobs.eligibility = request.POST["eligibility"]
        Jobs.job_description = request.POST["job_description"]
        Jobs.company_website = request.POST["company_website"]
        Jobs.job_location = request.POST["job_location"]
        
        Jobs.save()
        messages.success(request, "Job updated successfully.")
        return redirect("manage-job")
    
    return render(request, "update_jobs.html", {"jobs": Jobs})
    
    

from django.shortcuts import get_object_or_404, redirect
def delete_job(request, job_id):
    if "tpo_id" not in request.session:
        messages.error(request, "You need to log in first.")
        return redirect("tpo_login")
    
    tpo = Tpo.objects.get(id=request.session["tpo_id"])
    
    Jobs =  get_object_or_404(Job, id=job_id)
    print("you are deleting a job " ,job_id)
    Jobs.delete()
    messages.success(request, "Job deleted successfully.")
    return redirect("manage-job")

from django.shortcuts import render, get_object_or_404, redirect
from .models import Alumni
from django.contrib import messages

def add_alumni(request):
    if "tpo_id" not in request.session:
        messages.error(request, "You need to log in first.")
        return redirect("tpo_login")
    
    tpo = Tpo.objects.get(id=request.session["tpo_id"])
    if request.method == "POST":
        name = request.POST.get("name")
        email = request.POST.get("email")
        course = request.POST.get("course")
        passing_year = request.POST.get("passing_year")
        current_company = request.POST.get("current_company")
        position = request.POST.get("position")
        profile = request.FILES.get("profile_pic")

        Alumni.objects.create(
            name=name,
            email=email,
            course=course,
            passing_year=passing_year,
            current_company=current_company,
            position=position,
            profile_pic=profile
        )
        messages.success(request, "Alumni added successfully!")
        return redirect('alumni_list')

    return render(request, "add_alumni.html")


def update_alumni(request, alumni_id):
    if "tpo_id" not in request.session:
        messages.error(request, "You need to log in first.")
        return redirect("tpo_login")
    
    tpo = Tpo.objects.get(id=request.session["tpo_id"])
    alumni = get_object_or_404(Alumni, id=alumni_id)
    
    if request.method == "POST":
        alumni.name = request.POST.get("name")
        alumni.email = request.POST.get("email")
        alumni.course = request.POST.get("course")
        alumni.passing_year = request.POST.get("passing_year")
        alumni.current_company = request.POST.get("current_company")
        alumni.position = request.POST.get("position")
        alumni.save()
        messages.success(request, "Alumni updated successfully!")
        return redirect('alumni_list')

    return render(request, "update_alumni.html", {'alumni': alumni})


def toggle_alumni_visibility(request, alumni_id):
    if "tpo_id" not in request.session:
        messages.error(request, "You need to log in first.")
        return redirect("tpo_login")
    
    tpo = Tpo.objects.get(id=request.session["tpo_id"])
    alumni = get_object_or_404(Alumni, id=alumni_id)
    alumni.is_visible = not alumni.is_visible
    alumni.save()
    messages.success(request, f"Alumni visibility set to {'Visible' if alumni.is_visible else 'Hidden'}")
    return redirect('alumni_list')


def alumni_list(request):
    if "tpo_id" not in request.session:
        messages.error(request, "You need to log in first.")
        return redirect("tpo_login")
    
    tpo = Tpo.objects.get(id=request.session["tpo_id"])
    alumni = Alumni.objects.all().order_by("-id")
    return render(request, "alumni_list.html", {"alumni": alumni})


# from django.shortcuts import render, get_object_or_404, redirect
# from users.models import JobApplication
# from tpo.models import Job
# from users.models import User
# from users.models import APPLICATION_STATUS_CHOICES

# def tpo_applications_view(request):
#     applications = JobApplication.objects.select_related('student', 'job').all()

#     return render(request, 'tpo_applications.html', {
#         'applications': applications,
#         'APPLICATION_STATUS_CHOICES': APPLICATION_STATUS_CHOICES,  # If you need this in the template
#     })

# def update_application_status(request, application_id):
#     application = get_object_or_404(JobApplication, id=application_id)

#     if request.method == 'POST':
#         new_status = request.POST.get('status')
#         if new_status in dict(APPLICATION_STATUS_CHOICES):
#             application.status = new_status
#             application.save()
#             return redirect('tpo_applications_view')  # or wherever your dashboard is

#     return render(request, 'tpo/update_status.html', {
#         'application': application,
#         'status_choices': APPLICATION_STATUS_CHOICES
#     })
from django.shortcuts import render
from users.models import JobApplication

def tpo_applications_view(request):
    if "tpo_id" not in request.session:
        messages.error(request, "You need to log in first.")
        return redirect("tpo_login")
    
    tpo = Tpo.objects.get(id=request.session["tpo_id"])
    # Fetch all job applications for TPO to view
    applications = JobApplication.objects.select_related('student', 'job').all().order_by('-applied_at')
    
    # Pass the choices for the status field to the template
    status_choices = JobApplication._meta.get_field('status').choices
    
    return render(request, 'job_applications.html', {
        'applications': applications,
        'status_choices': status_choices,
    })

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from users.models import JobApplication

def update_application_status(request, application_id):
    if "tpo_id" not in request.session:
        messages.error(request, "You need to log in first.")
        return redirect("tpo_login")
    
    tpo = Tpo.objects.get(id=request.session["tpo_id"])
    # Fetch the application
    application = get_object_or_404(JobApplication, id=application_id)
    
    if request.method == 'POST':
        # Get the selected status from the form
        new_status = request.POST.get('status')
        if new_status:
            # Update the status
            application.status = new_status
            application.save()
            return redirect('tpo_applications')  # Redirect to the applications page

    # If it's not a POST request, return an error or a message
    return HttpResponse('Invalid request', status=400)

from django.shortcuts import render, redirect
from .models import Job, Drive, Tpo
from django.utils.dateparse import parse_datetime
from django.contrib import messages

def schedule_drive(request):
    if 'tpo_id' not in request.session:
        messages.error(request, "Please login first")
        return redirect('tpo_login')

    tpo = Tpo.objects.get(id=request.session['tpo_id'])

    if request.method == 'POST':
        job_id = request.POST.get('job_id')
        drive_date = parse_datetime(request.POST.get('drive_date'))
        venue = request.POST.get('venue')
        additional_info = request.POST.get('additional_info')

        job = Job.objects.get(id=job_id)

        Drive.objects.create(
            job=job,
            tpo=tpo,
            drive_date=drive_date,
            venue=venue,
            additional_info=additional_info
        )

        messages.success(request, "Drive scheduled successfully!")
        return redirect('tpo_dashboard')  # Redirect to TPO's dashboard or list

    jobs = Job.objects.all()  # Show all jobs to TPO
    return render(request, 'schedule_drive.html', {'jobs': jobs})


# from django.shortcuts import render, redirect
# from .models import Job, Drive, Tpo
# from django.utils.dateparse import parse_datetime
# from django.contrib import messages

# def schedule_drive(request):
#     if 'tpo_id' not in request.session:
#         messages.error(request, "Please login first")
#         return redirect('tpo_login')

#     tpo = Tpo.objects.get(id=request.session['tpo_id'])

#     if request.method == 'POST':
#         job_id = request.POST.get('job_id')
#         drive_date = parse_datetime(request.POST.get('drive_date'))
#         venue = request.POST.get('venue')
#         additional_info = request.POST.get('additional_info')

#         job = Job.objects.get(id=job_id)

#         Drive.objects.create(
#             job=job,
#             tpo=tpo,
#             drive_date=drive_date,
#             venue=venue,
#             additional_info=additional_info
#         )

#         messages.success(request, "Drive scheduled successfully!")
#         return redirect('tpo_dashboard')  # Redirect to TPO's dashboard or list

#     jobs = Job.objects.all()  # Show all jobs to TPO
#     return render(request, 'schedule_drive.html', {'jobs': jobs})

from django.shortcuts import render, get_object_or_404
from django.http import HttpResponseRedirect
from django.urls import reverse
from .models import Drive

def manage_drive(request):
    if "tpo_id" not in request.session:
        messages.error(request, "You need to log in first.")
        return redirect("tpo_login")
    
    tpo = Tpo.objects.get(id=request.session["tpo_id"])
    # Retrieve all drives from the database
    drives = Drive.objects.all()
    
    # Pass the drives to the template
    return render(request, 'manage_drive.html', {'drives': drives})


def update_drive(request, drive_id):
    if "tpo_id" not in request.session:
        messages.error(request, "You need to log in first.")
        return redirect("tpo_login")
    
    tpo = Tpo.objects.get(id=request.session["tpo_id"])
    # Retrieve the drive to update
    drive = get_object_or_404(Drive, id=drive_id)
    
    if request.method == 'POST':
        # Update drive fields with POST data (you can modify this based on your form fields)
        drive.drive_date = request.POST.get('drive_date')
        drive.venue = request.POST.get('venue')
        drive.additional_info = request.POST.get('additional_info')
        drive.save()
        
        # Redirect to the manage drive page after update
        return HttpResponseRedirect(reverse('manage-drive'))
    
    # Pass the drive to the template for updating
    return render(request, 'update_drive.html', {'drive': drive})

def update_drive_status(request, drive_id):
    if "tpo_id" not in request.session:
        messages.error(request, "You need to log in first.")
        return redirect("tpo_login")
    
    tpo = Tpo.objects.get(id=request.session["tpo_id"])
    drive = get_object_or_404(Drive, id=drive_id)
    new_status = request.POST.get("status")
    if new_status in ['ongoing', 'completed', 'cancelled']:
        drive.status = new_status
        drive.save()
        messages.success(request, f"Drive status updated to {new_status.capitalize()}.")
    else:
        messages.error(request, "Invalid status.")
    return redirect('manage-drive')

def delete_drive(request, drive_id):
    if "tpo_id" not in request.session:
        messages.error(request, "You need to log in first.")
        return redirect("tpo_login")
    
    tpo = Tpo.objects.get(id=request.session["tpo_id"])
    # Retrieve the drive to delete
    drive = get_object_or_404(Drive, id=drive_id)
    
    # Delete the drive
    drive.delete()
    
    # Redirect to the manage drive page after deletion
    return HttpResponseRedirect(reverse('manage-drive'))

from django.shortcuts import render
from .models import HiringPartner

def view_companies(request):
    if "tpo_id" not in request.session:
        messages.error(request, "You need to log in first.")
        return redirect("tpo_login")
    
    tpo = Tpo.objects.get(id=request.session["tpo_id"])
    companies = HiringPartner.objects.all()
    return render(request, 'view_companies.html', {'companies': companies})

# View to add a new Hiring Partner
def add_hiring_partner(request):
    if "tpo_id" not in request.session:
        messages.error(request, "You need to log in first.")
        return redirect("tpo_login")
    
    tpo = Tpo.objects.get(id=request.session["tpo_id"])
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        logo = request.FILES.get('logo')
        website = request.POST.get('website')
        year = request.POST.get('year', 2025)

        # Create a new HiringPartner
        HiringPartner.objects.create(
            name=name,
            description=description,
            logo=logo,
            website=website,
            year=year
        )

        return redirect('view_companies')

    return render(request, 'add_update_hiring_partner.html', {'action': 'Add'})

# View to update an existing Hiring Partner
def update_hiring_partner(request, pk):
    if "tpo_id" not in request.session:
        messages.error(request, "You need to log in first.")
        return redirect("tpo_login")
    
    tpo = Tpo.objects.get(id=request.session["tpo_id"])
    partner = get_object_or_404(HiringPartner, pk=pk)
    
    if request.method == 'POST':
        partner.name = request.POST.get('name')
        partner.description = request.POST.get('description')
        partner.logo = request.FILES.get('logo') or partner.logo  # Keep old logo if none is uploaded
        partner.website = request.POST.get('website')
        partner.year = request.POST.get('year', partner.year)

        partner.save()
        return redirect('view_companies')

    return render(request, 'add_update_hiring_partner.html', {'action': 'Update', 'partner': partner})

# View to delete an existing Hiring Partner
def delete_hiring_partner(request, pk):
    if "tpo_id" not in request.session:
        messages.error(request, "You need to log in first.")
        return redirect("tpo_login")
    
    tpo = Tpo.objects.get(id=request.session["tpo_id"])
    partner = get_object_or_404(HiringPartner, pk=pk)
    partner.delete()
    return redirect('view_companies')